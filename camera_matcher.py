"""
Модуль сопоставления комплексов фотовидеофиксации с очагами ДТП.

Поддерживаемые форматы адреса камер (название дороги + пикетаж в одной ячейке):
  - ФАД Р-217 «Кавказ» 775км +890м (справа), ...
  - а/д «Магарамкент-Ахты-Рутул» 25км +500м, ...
  - а/д «Махачкала-Турали-Каспийск» г. Махачкала, пр-т. ...
  - г. Махачкала, перекресток ул. ... (городские, без пикетажа)

Стратегии сопоставления:
  1. Линейный очаг с пикетажем → камера на том же участке дороги (по пикетажу)
  2. Линейный очаг без камеры на участке → гео-поиск от крайних точек (1 км НП / 500 м вне НП)
  3. Очаг без пикетажа (НП) → гео-поиск 200 м (закрытый) и 500 м (ближайшие)
"""

import json
import logging
import math
import os
import re
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ========================
# Константы
# ========================

# Кэш камер (в постоянном хранилище, /data на Amvera)
_DATA_DIR = os.environ.get("CAMERA_DATA_DIR", "data")
CAMERA_CACHE_DIR = os.path.join(_DATA_DIR, "cameras")

# Радиусы поиска (метры)
RADIUS_IN_SETTLEMENT_WITHOUT_PIKETAZH = 200    # Очаг без пикетажа в НП — «закрытый»
RADIUS_IN_SETTLEMENT_NEARBY = 500              # Очаг без пикетажа в НП — «ближайшие»
RADIUS_NEAR_ENDPOINT_IN_NP = 1000              # Линейный очаг в НП — поиск от крайних точек
RADIUS_NEAR_ENDPOINT_OUTSIDE_NP = 500          # Линейный очаг вне НП — поиск от крайних точек

# Допуск пикетажа (метры) для прямого совпадения
PIKETAZH_TOLERANCE_M = 100


# ========================
# Структура данных камеры
# ========================

# Каждая камера — dict с ключами:
#   id: str          — ID из файла
#   number: str      — номер комплекса (столбец «Комплекс»)
#   model: str       — модель (СКАТ-С, Вокорд и т.д.)
#   lat: float       — широта
#   lon: float       — долгота
#   address: str     — полный адрес (сырой)
#   road_name: str   — извлечённое название дороги (нормализованное)
#   road_number: str | None — номер трассы (Р-217, А-167) если есть
#   km: float | None — километры (из пикетажа)
#   m: float | None  — метры (из пикетажа)
#   piketazh_m: float | None — пикетаж в метрах от начала дороги


# ========================
# Кэш камер
# ========================

_camera_cache: dict[str, list[dict]] = {}  # region_code → [camera, ...]


def _ensure_cache_dir():
    os.makedirs(CAMERA_CACHE_DIR, exist_ok=True)


def _cache_path(region_code: str) -> str:
    return os.path.join(CAMERA_CACHE_DIR, f"cameras_{region_code}.json")


# ========================
# Парсинг файла камер
# ========================

def parse_camera_file(file_bytes: bytes) -> list[dict]:
    """
    Парсит Excel-файл камер ГИБДД.РФ.

    Формат файла:
      Строка 1: заголовок (пропускается)
      Строка 2: подзаголовки столбцов (#, ID, Комплекс, Модель, Координаты, Адрес, ...)
      Строка 3: подзаголовки координат (Широта, Долгота)
      Строка 4: пустая / номера столбцов
      Строка 5+: данные

    Возвращает список dict с данными камер.
    """
    cameras = []
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Не удалось открыть файл камер: {e}")
        return cameras

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(min_row=5, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue

            try:
                camera_id = str(row[1]).strip() if row[1] else ""
                number = str(row[2]).strip() if row[2] else ""
                model = str(row[3]).strip() if row[3] else ""
                lat = _parse_float(row[4])
                lon = _parse_float(row[5])
                address = str(row[6]).strip() if row[6] else ""

                if lat is None or lon is None:
                    continue

                # Извлекаем дорогу и пикетаж из адреса
                road_name, road_number, km, m_val = parse_camera_address(address)

                piketazh_m = None
                if km is not None:
                    piketazh_m = km * 1000 + (m_val or 0)

                cameras.append({
                    "id": camera_id,
                    "number": number,
                    "model": model,
                    "lat": lat,
                    "lon": lon,
                    "address": address,
                    "road_name": road_name,
                    "road_number": road_number,
                    "km": km,
                    "m": m_val,
                    "piketazh_m": piketazh_m,
                })
            except Exception as e:
                logger.warning(f"Ошибка парсинга строки камеры: {e}")
                continue

    wb.close()
    logger.info(f"Парсинг камер: {len(cameras)} камер из файла")
    return cameras


def parse_camera_address(address: str) -> tuple[str, str | None, float | None, float | None]:
    """
    Извлекает название дороги, номер трассы и пикетаж из адреса камеры.

    Возвращает: (road_name, road_number, km, m)
      - road_name: нормализованное название дороги (lowercase, без лишних пробелов)
      - road_number: номер трассы (Р-217, А-167) или None
      - km: километры (float) или None
      - m: метры (float) или None

    Форматы:
      ФАД Р-217 «Кавказ» 775км +890м (справа), ...
      а/д «Магарамкент-Ахты-Рутул» 25км +500м, ...
      а/д «Махачкала-Буйнакск-Леваши-В.Гуниб» 0км +500м (справа), ...
      ФАД «Кавказ»-Шамхал-Красноармейское 12км +000м, ...
      а/д «Махачкала-Турали-Каспийск» г. Махачкала, пр-т. ... (без пикетажа)
      г. Махачкала, перекресток ул. ... (городские, без пикетажа)
    """
    road_name = ""
    road_number = None
    km = None
    m_val = None

    # 1. Извлекаем пикетаж: шаблоны «775км +890м», «0км +500м», «12км +000м»
    piket_match = re.search(
        r'(\d+)\s*км\s*\+\s*(\d+)\s*м',
        address, re.IGNORECASE
    )
    if piket_match:
        km = float(piket_match.group(1))
        m_val = float(piket_match.group(2))

    # 2. Извлекаем номер трассы: Р-217, А-167, М-4, Р-275 и т.д.
    road_num_match = re.search(
        r'\b([РАМ]-\d+(?:\s*\(.*?\))?)\b',
        address
    )
    if road_num_match:
        raw_num = road_num_match.group(1)
        # Очищаем от скобок для мэтчинга: Р-217(основное) → Р-217
        road_number = re.sub(r'\s*\(.*\)', '', raw_num).strip()

    # 3. Извлекаем название дороги из кавычек «...» или "...»
    #    Формат: «Кавказ», «Магарамкент-Ахты-Рутул», «Махачкала-Буйнакск-...»
    quoted_match = re.search(r'[«""](.+?)[»""]', address)
    if quoted_match:
        road_name = _normalize_road_name(quoted_match.group(1))

    # 4. Для ФАД без кавычек у названия: ФАД-Башлыкент, ФАД «Кавказ»-Шамхал-...
    #    Если название в кавычках — короткое (одно слово, типа «Кавказ»),
    #    а после кавычек идёт тире и ещё название дороги
    if road_name and len(road_name.split('-')) == 1 and road_number:
        # Возможно, после кавычек есть продолжение: «Кавказ»-Шамхал-Красноармейское
        after_quote = address[quoted_match.end():] if quoted_match else ""
        ext_match = re.match(r'\s*[-—]\s*([А-Яа-яЁё][\w\-]+(?:\s*[-—]\s*[А-Яа-яЁё][\w\-]+)*)', after_quote)
        if ext_match:
            road_name = _normalize_road_name(ext_match.group(1))

    # 5. Для формата без кавычек: ФАД-Башлыкент, кизляр-н.бирюзак
    if not road_name:
        # Формат: а/д <название> или ФАД-<название>
        no_quote_match = re.search(
            r'(?:а/д|ФАД)\s*[«""]?(.+?)(?:\s+\d+км|\s*,\s*(?:г\.|возле|вдоль|с\.|Дагестан))',
            address
        )
        if no_quote_match:
            road_name = _normalize_road_name(no_quote_match.group(1).strip(' «"»'))

    return road_name, road_number, km, m_val


def _normalize_road_name(name: str) -> str:
    """
    Нормализует название дороги для сравнения:
    - нижний регистр
    - удаляет лишние пробелы вокруг тире
    - убирает тип дороги (ФАД, а/д, автомобильная дорога)
    - убирает пояснения в скобках (основное направление и т.п.)
    """
    s = name.strip().lower()
    # Убираем пояснения в скобках
    s = re.sub(r'\(.*?\)', '', s)
    # Убираем типы дорог
    s = re.sub(r'^(федеральная\s+)?автомобильная\s+дорога\s*', '', s)
    s = re.sub(r'^(фад|а/д|а\.д\.)\s*', '', s)
    # Нормализуем тире: « - » → «-»
    s = re.sub(r'\s*[-—]\s*', '-', s)
    # Убираем множественные пробелы
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


# ========================
# Гео-утилиты
# ========================

def haversine(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Расстояние между двумя точками в метрах (формула Гаверсинуса)."""
    R = 6_371_000  # радиус Земли в метрах
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = (
        math.sin(dphi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(dlambda / 2) ** 2
    )
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ========================
# Нормализация названий дорог ГИБДД
# ========================

def normalize_gibdd_road(road: str) -> tuple[str, str | None]:
    """
    Нормализует название дороги из данных ГИБДД (поле `dor`).

    Возвращает: (normalized_name, road_number)
      - normalized_name: для сравнения с камерами (lowercase, без лишнего)
      - road_number: номер трассы (Р-217, А-167) или None
    """
    if not road:
        return "", None

    # Извлекаем номер трассы
    num_match = re.search(r'\b([РАМ]-\d+)\b', road)
    road_number = num_match.group(1) if num_match else None

    # Нормализуем полное название
    normalized = _normalize_road_name(road)
    return normalized, road_number


# ========================
# Мэтчинг: дорога очага ↔ дорога камеры
# ========================

def _roads_match(
    cluster_road: str,
    cluster_road_number: str | None,
    camera_road_name: str,
    camera_road_number: str | None,
) -> bool:
    """
    Проверяет совпадение дороги очага и дороги камеры.

    Стратегии (по приоритету):
    1. Совпадение номера трассы (Р-217 == Р-217)
    2. Нормализованные названия совпадают (магарамкент-ахты-рутул == магарамкент-ахты-рутул)
    """
    # Стратегия 1: по номеру трассы
    if cluster_road_number and camera_road_number:
        if cluster_road_number.upper() == camera_road_number.upper():
            return True

    # Стратегия 2: по нормализованному названию
    if cluster_road and camera_road_name:
        norm_cluster = _normalize_road_name(cluster_road)
        norm_camera = camera_road_name
        if not norm_cluster or not norm_camera:
            return False

        # Прямое совпадение
        if norm_cluster == norm_camera:
            return True

        # Частичное: одно название содержится в другом
        # (на случай, если одно короче другого из-за обрезки)
        if norm_cluster in norm_camera or norm_camera in norm_cluster:
            return True

    return False


# ========================
# Основная функция мэтчинга
# ========================

def match_cameras_to_clusters(
    clusters: list[dict],
    cameras: list[dict],
) -> list[dict]:
    """
    Сопоставляет камеры с очагами ДТП по алгоритму:

    1. Линейные очаги с пикетажем:
       а) Ищем камеру на том же участке дороги (по пикетажу) → «закрытый»
       б) Если нет — ищем от крайних точек: 1 км в НП / 500 м вне НП
    2. Очаги без пикетажа (НП-перекрёстки, НП-участки):
       а) Камера в радиусе 200 м → «закрытый»
       б) Камера в радиусе 500 м → «ближайшая»

    Каждому кластеру добавляются ключи:
      camera_in_cluster: dict | None  — камера прямо в очаге
      nearest_cameras: list[dict]     — ближайшие камеры (вне очага)

    Args:
        clusters: список кластеров (очагов) из concentration_points
        cameras: список камер из parse_camera_file

    Returns:
        тот же список clusters, обогащённый данными о камерах
    """
    if not cameras:
        logger.info("Камеры не загружены — пропуск сопоставления")
        for cluster in clusters:
            cluster["camera_in_cluster"] = None
            cluster["nearest_cameras"] = []
        return clusters

    logger.info(
        f"Сопоставление камер: {len(clusters)} очагов × {len(cameras)} камер"
    )

    for cluster in clusters:
        zone_type = cluster.get("zone_type", "")
        has_piketazh = cluster.get("has_piketazh", False)
        cluster_road = cluster.get("road", "")
        _, cluster_road_num = normalize_gibdd_road(cluster_road)

        # Предфильтруем камеры по совпадению дороги (если есть)
        if cluster_road:
            road_cameras = [
                c for c in cameras
                if _roads_match(cluster_road, cluster_road_num,
                                c["road_name"], c["road_number"])
            ]
        else:
            road_cameras = []

        # Координаты очага
        first_coords = cluster.get("first_coords")
        last_coords = cluster.get("last_coords")
        start_km = cluster.get("start_km")
        end_km = cluster.get("end_km")

        camera_in_cluster = None
        nearest_cameras = []

        # ---- Линейный очаг с пикетажем ----
        if has_piketazh and start_km is not None and end_km is not None:
            # 1а) Ищем камеру на участке дороги по пикетажу
            camera_in_cluster = _find_camera_by_piketazh(
                road_cameras, start_km, end_km
            )

            # 1б) Если нет — ищем от крайних точек по координатам
            if not camera_in_cluster and first_coords and last_coords:
                is_np = zone_type.startswith("settlement")
                radius = RADIUS_NEAR_ENDPOINT_IN_NP if is_np else RADIUS_NEAR_ENDPOINT_OUTSIDE_NP

                # Ищем от обеих крайних точек
                endpoints = [first_coords, last_coords]
                nearest_cameras = _find_nearest_cameras(
                    cameras, endpoints, radius, exclude_cluster=True
                )

        # ---- Очаг без пикетажа (НП) ----
        elif zone_type.startswith("settlement") and first_coords:
            # 2а) Камера в 200 м → закрытый
            nearest_cameras_200 = _find_nearest_cameras(
                cameras, [first_coords], RADIUS_IN_SETTLEMENT_WITHOUT_PIKETAZH
            )
            if nearest_cameras_200:
                camera_in_cluster = nearest_cameras_200[0]

            # 2б) Камеры в 500 м → ближайшие
            nearest_cameras = _find_nearest_cameras(
                cameras, [first_coords], RADIUS_IN_SETTLEMENT_NEARBY
            )

        # ---- Вне НП без пикетажа (редкий случай) ----
        elif not zone_type.startswith("settlement") and first_coords:
            nearest_cameras = _find_nearest_cameras(
                cameras, [first_coords], RADIUS_NEAR_ENDPOINT_OUTSIDE_NP
            )

        cluster["camera_in_cluster"] = camera_in_cluster
        cluster["nearest_cameras"] = nearest_cameras

    # Статистика
    covered = sum(1 for c in clusters if c["camera_in_cluster"])
    has_nearby = sum(1 for c in clusters if c["nearest_cameras"])
    logger.info(
        f"Результат сопоставления: "
        f"{covered}/{len(clusters)} очагов закрыты камерами, "
        f"{has_nearby}/{len(clusters)} имеют ближайшие камеры"
    )

    return clusters


# ========================
# Вспомогательные функции мэтчинга
# ========================

def _find_camera_by_piketazh(
    road_cameras: list[dict],
    start_km: float,
    end_km: float,
) -> dict | None:
    """
    Ищет камеру на участке дороги по пикетажу.

    Камера считается на участке, если её piketazh_m попадает
    в диапазон [start_km*1000 - TOLERANCE, end_km*1000 + TOLERANCE].

    Если камер несколько — берём ближайшую к середине участка.
    """
    if not road_cameras:
        return None

    start_m = start_km * 1000
    end_m = end_km * 1000
    mid_m = (start_m + end_m) / 2

    candidates = []
    for cam in road_cameras:
        pik_m = cam.get("piketazh_m")
        if pik_m is None:
            continue
        if start_m - PIKETAZH_TOLERANCE_M <= pik_m <= end_m + PIKETAZH_TOLERANCE_M:
            dist_to_mid = abs(pik_m - mid_m)
            candidates.append((dist_to_mid, cam))

    if not candidates:
        return None

    # Сортируем по расстоянию до середины участка
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


def _find_nearest_cameras(
    cameras: list[dict],
    reference_points: list[tuple[float, float]],
    max_radius_m: float,
    exclude_cluster: bool = False,
) -> list[dict]:
    """
    Ищет камеры вблизи заданных точек.

    Args:
        cameras: все камеры
        reference_points: список (lat, lon) опорных точек
        max_radius_m: максимальный радиус поиска
        exclude_cluster: если True, исключаем камеры, уже попавшие в очаг
                         (по пикетажу на той же дороге)

    Returns:
        Список dict с добавленными ключами "distance_m" и "ref_point_idx",
        отсортированный по расстоянию. Не более 3 камер.
    """
    results = []

    for cam in cameras:
        cam_lat = cam["lat"]
        cam_lon = cam["lon"]

        min_dist = float("inf")
        best_idx = 0

        for idx, (lat, lon) in enumerate(reference_points):
            d = haversine(lat, lon, cam_lat, cam_lon)
            if d < min_dist:
                min_dist = d
                best_idx = idx

        if min_dist <= max_radius_m:
            results.append({
                **cam,
                "distance_m": round(min_dist, 1),
                "ref_point_idx": best_idx,
            })

    # Сортируем по расстоянию, берём топ-3
    results.sort(key=lambda x: x["distance_m"])
    return results[:3]


# ========================
# Управление кэшем камер
# ========================

def save_cameras(region_code: str, cameras: list[dict]) -> None:
    """Сохраняет камеры для региона в память + на диск."""
    _ensure_cache_dir()
    _camera_cache[region_code] = cameras

    # Сохраняем на диск (без лишних данных)
    cache_data = []
    for cam in cameras:
        cache_data.append({
            "id": cam["id"],
            "number": cam["number"],
            "model": cam["model"],
            "lat": cam["lat"],
            "lon": cam["lon"],
            "address": cam["address"],
            "road_name": cam["road_name"],
            "road_number": cam["road_number"],
            "km": cam["km"],
            "m": cam["m"],
            "piketazh_m": cam["piketazh_m"],
        })

    path = _cache_path(region_code)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2)
    logger.info(
        f"Камеры для региона {region_code}: сохранено {len(cameras)} "
        f"в кэш ({path})"
    )


def load_cameras(region_code: str) -> list[dict] | None:
    """
    Загружает камеры для региона из памяти или с диска.

    Returns list[dict] или None, если камер для региона нет.
    """
    # Из памяти
    if region_code in _camera_cache:
        return _camera_cache[region_code]

    # С диска
    path = _cache_path(region_code)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                cameras = json.load(f)
            _camera_cache[region_code] = cameras
            logger.info(
                f"Камеры для региона {region_code}: "
                f"загружено {len(cameras)} из кэша"
            )
            return cameras
        except Exception as e:
            logger.error(f"Ошибка загрузки камер из кэша: {e}")

    return None


def get_cached_regions() -> list[str]:
    """Возвращает список кодов регионов, для которых есть камеры."""
    _ensure_cache_dir()
    regions = []
    if os.path.exists(CAMERA_CACHE_DIR):
        for f in os.listdir(CAMERA_CACHE_DIR):
            if f.startswith("cameras_") and f.endswith(".json"):
                region = f.replace("cameras_", "").replace(".json", "")
                regions.append(region)
    return sorted(regions)


# ========================
# Утилиты
# ========================

def _parse_float(val) -> float | None:
    """Парсит float из значения ячейки."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None