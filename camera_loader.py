"""
Модуль загрузки, парсинга и поиска камер фотовидеофиксации.

Формат файла: Excel (gibddrf_cameras_change_*.xlsx)
Лист: cameras_list_report
Строки 1-3: заголовки, данные с 5-й строки.

Структура камеры после парсинга:
    {
        "id": "820000000000000665",
        "number": "1811003",
        "model": "СКАТ-С",
        "lat": 43.12228,
        "lon": 47.087997,
        "address": "ФАД Р-217 «Кавказ» 775км +890м ...",
        "road_num": "Р-217",        # номер дороги (Р-217, А-167) или None
        "road_name": "Кавказ",       # название в кавычках или None
        "road_simple": "кавказ",     # нормализованное для мэтчинга
        "piket": 775.890,            # пикетаж в км (None если нет)
        "has_piket": True,
    }
"""

import io
import html as html_mod
import logging
import math
import re
import xml.etree.ElementTree as ET
from typing import Optional

import openpyxl
try:
    import xlrd
except ImportError:
    xlrd = None  # type: ignore[assignment]


logger = logging.getLogger(__name__)

# ========================
# Регулярные выражения
# ========================

# Номер дороги: Р-217, А-167, М-4 (с опциональным пробелом после буквы)
_RE_ROAD_NUM = re.compile(r'([РАМ]-?\s*\d+)', re.IGNORECASE)

# Пикетаж: "775км +890м", "82км +080м", "0км +000м"
_RE_PIKET = re.compile(r'(\d+)\s*км\s*\+\s*(\d+)\s*м?')

# Название дороги в кавычках (елочки и стандартные)
_RE_QUOTED = re.compile(r'[«\"\u201c](.+?)[»\"\u201d]')

# Городская камера (перекрёсток, улица)
_RE_URBAN = re.compile(
    r'(перекресток|пер\.\s|ул\.|пр-т|пр-кт|проспект|просп\.)',
    re.IGNORECASE,
)


# ========================
# Парсинг
# ========================

def parse_camera_file(file_bytes: bytes) -> list[dict]:
    """
    Парсит Excel-файл со списком камер (.xlsx, .xls или XML Spreadsheet).
    """
    logger.info(f"parse_camera_file: {len(file_bytes)} байт, "
                f"начало: {file_bytes[:20]!r}")

    if not file_bytes or len(file_bytes) < 4:
        logger.error("parse_camera_file: пустой или слишком маленький файл")
        return []

    # Определяем формат по сигнатуре
    zip_sig = bytes([0x50, 0x4B, 0x03, 0x04])
    ole_sig = bytes([0xD0, 0xCF, 0x11, 0xE0, 0xA1, 0xB1, 0x1A, 0xE1])
    is_xml = file_bytes[:5] == b'<?xml'

    if file_bytes[:4] == zip_sig:
        logger.info("parse_camera_file: формат .xlsx (ZIP)")
        return _parse_xlsx(file_bytes)
    elif file_bytes[:8] == ole_sig:
        logger.info("parse_camera_file: формат .xls (OLE), используем xlrd")
        return _parse_xls(file_bytes)
    elif is_xml:
        logger.info("parse_camera_file: формат XML Spreadsheet")
        return _parse_xml(file_bytes)
    else:
        sig = file_bytes[:8].hex()
        logger.error(f"parse_camera_file: неизвестный формат, сигнатура: {sig}")
        raise ValueError(f"Неизвестный формат файла (сигнатура: {sig})")


def _row_to_camera(row, xml_mode: bool = False) -> dict | None:
    """Общая логика извлечения камеры из строки.

    Стандартный формат (xlsx/xls через openpyxl/xlrd, данные с row 5):
        0=№, 1=ID, 2=Комплекс, 3=Модель, 4=Широта, 5=Долгота, 6=Адрес

    XML Spreadsheet формат (данные с row 4, 21 столбец):
        0=№, 1=ID, 4=Комплекс, 7=Модель, 9=Широта, 10=Долгота, 11=Адрес
    """
    if not row:
        return None

    if xml_mode:
        # XML: нужно минимум 12 столбцов
        if len(row) < 12:
            return None
        if not row[0]:
            return None
        col_num, col_id, col_complex, col_model = 0, 1, 4, 7
        col_lat, col_lon, col_addr = 9, 10, 11
    else:
        # Стандартный xlsx/xls
        if len(row) < 7:
            return None
        if not row[0] or row[2] is None:
            return None
        col_num, col_id, col_complex, col_model = 0, 1, 2, 3
        col_lat, col_lon, col_addr = 4, 5, 6

    try:
        lat = float(row[col_lat]) if row[col_lat] else None
        lon = float(row[col_lon]) if row[col_lon] else None
    except (ValueError, TypeError):
        lat = lon = None

    if lat is None or lon is None:
        return None

    address = str(row[col_addr]).strip() if row[col_addr] else ""

    road_num, road_name, road_simple, piket, has_piket = (
        _parse_camera_address(address)
    )

    return {
        "id": str(row[col_id]) if len(row) > col_id and row[col_id] else "",
        "number": str(row[col_complex]).strip() if len(row) > col_complex and row[col_complex] else "",
        "model": str(row[col_model]).strip() if len(row) > col_model and row[col_model] else "",
        "lat": lat,
        "lon": lon,
        "address": address,
        "road_num": road_num,
        "road_name": road_name,
        "road_simple": road_simple,
        "piket": piket,
        "has_piket": has_piket,
    }


def _parse_xls(file_bytes: bytes) -> list[dict]:
    """Парсинг .xls файла через xlrd."""
    if xlrd is None:
        raise ImportError(
            "xlrd не установлен. Выполните: pip install xlrd"
        )
    wb = xlrd.open_workbook(file_contents=file_bytes)
    ws = wb.sheet_by_index(0)

    cameras = []
    for row_idx in range(4, ws.nrows):  # данные с 5-й строки (индекс 4)
        row = ws.row_values(row_idx)
        cam = _row_to_camera(row)
        if cam:
            cameras.append(cam)

    wb.release_resources()
    _log_result(cameras)
    return cameras


def _parse_xlsx(file_bytes: bytes) -> list[dict]:
    """Парсинг .xlsx файла через openpyxl."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]

    cameras = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        cam = _row_to_camera(row)
        if cam:
            cameras.append(cam)

    wb.close()
    _log_result(cameras)
    return cameras


# HTML-сущности, которые часто встречаются в XML Spreadsheet от Excel,
# но не являются стандартными XML-сущностями
_HTML_ENTITY_MAP = {
    "laquo": "\u00ab",    # «
    "raquo": "\u00bb",    # »
    "ldquo": "\u201c",    # "
    "rdquo": "\u201d",    # "
    "lsquo": "\u2018",    # '
    "rsquo": "\u2019",    # '
    "ndash": "\u2013",    # –
    "mdash": "\u2014",    # —
    "nbsp": "\u00a0",     # неразрывный пробел
    "quot": '"',
    "amp": "&",
    "lt": "<",
    "gt": ">",
}

_HTML_ENTITY_RE = re.compile(r"&(\w+);")


def _replace_html_entities(text: str) -> str:
    """Заменяет HTML-сущности на Unicode-символы. Неизвестные — удаляет."""
    def _repl(m):
        name = m.group(1)
        return _HTML_ENTITY_MAP.get(name, "")
    return _HTML_ENTITY_RE.sub(_repl, text)


def _parse_xml(file_bytes: bytes) -> list[dict]:
    """Парсинг XML Spreadsheet (SpreadsheetML) — файл .xls с '<?xml' внутри."""
    # Заменяем HTML-сущности, которые стандартный XML-парсер не знает
    text = file_bytes.decode("utf-8", errors="replace")
    text = _replace_html_entities(text)
    xml_bytes = text.encode("utf-8")

    try:
        root = ET.fromstring(xml_bytes)
    except ET.ParseError as e:
        # Пробуем убрать BOM
        logger.warning(f"XML parse error, пробуем без BOM: {e}")
        cleaned = xml_bytes.lstrip(b'\xef\xbb\xbf')
        root = ET.fromstring(cleaned)

    # Ищем первый Worksheet
    ws = root.find(".//{urn:schemas-microsoft-com:office:spreadsheet}Worksheet")
    if ws is None:
        # Пробуем без namespace
        ws = root.find(".//Worksheet")
    if ws is None:
        raise ValueError("Не найден Worksheet в XML")

    # Собираем строки таблицы
    rows = ws.findall(".//{urn:schemas-microsoft-com:office:spreadsheet}Row")
    if not rows:
        rows = ws.findall(".//Row")

    cameras = []
    for row_idx, row_el in enumerate(rows):
        try:
            cells = row_el.findall(".//{urn:schemas-microsoft-com:office:spreadsheet}Cell")
            if not cells:
                cells = row_el.findall(".//Cell")

            # Ячейки могут идти с пропуском (ss:Index). Восстанавливаем порядок.
            row_values = []
            for cell in cells:
                idx_attr = cell.get(
                    "{urn:schemas-microsoft-com:office:spreadsheet}Index"
                ) or cell.get("Index")
                if idx_attr:
                    idx = int(idx_attr) - 1  # 1-based → 0-based
                    while len(row_values) < idx:
                        row_values.append(None)
                data_el = cell.find(
                    "{urn:schemas-microsoft-com:office:spreadsheet}Data"
                )
                if data_el is None:
                    data_el = cell.find("Data")
                value = data_el.text if data_el is not None and data_el.text else None
                row_values.append(value)

            cam = _row_to_camera(row_values, xml_mode=True)
            if cam:
                cameras.append(cam)
        except Exception as e:
            logger.warning(f"XML row[{row_idx}] error: {e}, row_values len={len(row_values) if 'row_values' in dir() else '?'}")

    _log_result(cameras)
    return cameras


def _log_result(cameras: list[dict]) -> None:
    with_piket = sum(1 for c in cameras if c["has_piket"])
    logger.info(
        f"Загружено {len(cameras)} камер "
        f"(с пикетажем: {with_piket}, городских: {len(cameras) - with_piket})"
    )


def _parse_camera_address(
    address: str,
) -> tuple[Optional[str], Optional[str], Optional[str], Optional[float], bool]:
    """
    Извлекает из адреса камеры: номер дороги, название, пикетаж.

    Returns:
        (road_num, road_name, road_simple, piket_km, has_piket)
    """
    road_num = None
    road_name = None
    road_simple = None
    piket = None
    has_piket = False

    # 1. Номер дороги
    m_num = _RE_ROAD_NUM.search(address)
    if m_num:
        road_num = m_num.group(1).upper().replace(" ", "")

    # 2. Название в кавычках
    m_quoted = _RE_QUOTED.search(address)
    if m_quoted:
        road_name = m_quoted.group(1)
        road_simple = road_name.lower().replace("-", " ")
        road_simple = re.sub(r"\s+", " ", road_simple).strip()

    # 3. Пикетаж
    m_piket = _RE_PIKET.search(address)
    if m_piket:
        has_piket = True
        piket = int(m_piket.group(1)) + int(m_piket.group(2)) / 1000.0

    return road_num, road_name, road_simple, piket, has_piket


# ========================
# Нормализация дорог ГИБДД
# ========================

def normalize_gibdd_road(
    dor: str,
) -> tuple[Optional[str], Optional[str]]:
    """
    Нормализует название дороги из карточки ДТП.

    Returns:
        (road_num, road_simple)
        - road_num: "Р-217", "А-167" или None
        - road_simple: "р 217 кавказ", "махачкала буйнакск леваши в.гуниб"
    """
    if not dor:
        return None, None

    d = dor.strip()

    # Номер дороги
    m = _RE_ROAD_NUM.search(d)
    road_num = m.group(1).upper().replace(" ", "") if m else None

    # Убираем всё после скобки "(основное направление)"
    d_clean = re.split(r"\s*\(", d)[0].strip()
    # Убираем "автомобильная дорога"
    d_clean = re.sub(
        r"автомобильная дорога\s*", "", d_clean, flags=re.IGNORECASE
    )
    # Нормализуем
    simple = d_clean.lower().replace("«", "").replace("»", "").replace('"', "")
    simple = simple.replace("-", " ")
    simple = re.sub(r"\s+", " ", simple).strip()

    return road_num, simple


# ========================
# Мэтчинг дорог
# ========================

def roads_match(gibdd_dor: str, camera: dict) -> bool:
    """
    Проверяет, относится ли камера к той же дороге, что и ДТП.

    Мэтчинг по приоритету:
    1. Совпадение номера дороги (Р-217 == Р-217)
    2. Вхождение нормализованных названий друг в друга
    3. Пересечение ключевых слов (>50% совпадение)
    """
    g_num, g_simple = normalize_gibdd_road(gibdd_dor)

    if not g_num and not g_simple:
        return False

    c_num = camera.get("road_num")
    c_simple = camera.get("road_simple")

    # 1. По номеру дороги
    if g_num and c_num and g_num == c_num:
        return True

    # 2. По нормализованным названиям (вхождение)
    if g_simple and c_simple:
        if g_simple in c_simple or c_simple in g_simple:
            return True
        # 3. Пересечение слов
        if _words_overlap(g_simple, c_simple, threshold=0.4):
            return True

    return False


def _words_overlap(a: str, b: str, threshold: float = 0.5) -> bool:
    """Доля общих слов относительно большего множества."""
    words_a = set(a.split())
    words_b = set(b.split())
    if not words_a or not words_b:
        return False
    common = words_a & words_b
    score = len(common) / max(len(words_a), len(words_b))
    return score >= threshold


# ========================
# Расстояние (Haversine)
# ========================

def haversine_km(lat1: float, lon1: float, lat2: float, lon2: float) -> float:
    """Расстояние в километрах между двумя точками."""
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(dlon / 2) ** 2
    )
    return R * 2 * math.asin(math.sqrt(a))


# ========================
# Поиск камер для очагов
# ========================

# Радиусы поиска (в км)
CLOSE_RADIUS_KM = 0.2         # 200 м — «закрытый» для очагов без пикетажа
NEAR_RADIUS_KM = 0.5           # 500 м — «ближайшие» для очагов без пикетажа
NEAR_ROAD_PK_NP_KM = 0.5     # 500 м по пикетажу — ближайшие в НП
NEAR_ROAD_PK_OUT_KM = 1.0     # 1 км по пикетажу — ближайшие вне НП


def find_cameras_for_cluster(
    cluster: dict,
    cameras: list[dict],
) -> dict:
    """
    Ищет камеры для одного очага концентрации ДТП.

    Логика:
    1. Очаг с пикетажем + дорогой:
       а) Камера на той же дороге, пикетаж в диапазоне реальных границ ДТП
          (dtp_pk_min .. dtp_pk_max) → «закрытый»
       б) Камера на той же дороге, пикетаж в пределах 500 м (в НП)
          или 1 км (вне НП) от границ ДТП → «ближайшая»
    2. Очаг без пикетажа (обычно НП):
       а) GPS-радиус 200 м → «закрытый»
       б) GPS-радиус 500 м → «ближайшая»

    Args:
        cluster: Словарь очага из concentration_points.
        cameras: Список камер (от parse_camera_file).

    Returns:
        {
            "in_cluster": camera_dict | None,   # камера в очаге
            "nearest": camera_dict | None,       # ближайшая камера
            "nearest_dist_m": int | None,        # расстояние до ближайшей (м)
            "status": "закрыт" | "открыт",
        }
    """
    result = {
        "in_cluster": None,
        "nearest": None,
        "nearest_dist_m": None,
        "status": "открыт",
    }

    if not cameras:
        return result

    road_name = cluster.get("road", "")
    zone_type = cluster.get("zone_type", "")
    start_piket = cluster.get("start_km")   # float км — окно группировки
    end_piket = cluster.get("end_km")       # float км — окно группировки
    dtp_pk_min = cluster.get("dtp_pk_min")  # реальный мин. пикетаж ДТП
    dtp_pk_max = cluster.get("dtp_pk_max")  # реальный макс. пикетаж ДТП
    first_coords = cluster.get("first_coords")  # (lat, lon) или None
    last_coords = cluster.get("last_coords")    # (lat, lon) или None

    has_piket = cluster.get("has_piketazh", False)
    if start_piket is not None and end_piket is not None:
        has_piket = True
    is_settlement = zone_type.startswith("settlement")

    if has_piket and road_name:
        # === Сценарий 1: Очаг с пикетажем ===
        result = _search_by_piketage(
            road_name, start_piket, end_piket,
            dtp_pk_min, dtp_pk_max,
            is_settlement, cameras,
        )
    else:
        # === Сценарий 2: Очаг без пикетажа (GPS) ===
        result = _search_by_gps(
            first_coords, last_coords, cameras,
        )

    return result


def _search_by_piketage(
    road_name: str,
    start_piket: float,
    end_piket: float,
    dtp_pk_min: float | None,
    dtp_pk_max: float | None,
    is_settlement: bool,
    cameras: list[dict],
) -> dict:
    """Поиск камер по пикетажу и названию дороги.

    Для определения «закрыт» используем реальные границы ДТП (dtp_pk_min/dtp_pk_max).
    Для поиска «ближайших» используем окно группировки (start_piket/end_piket).
    """
    result = {
        "in_cluster": None,
        "nearest": None,
        "nearest_dist_m": None,
        "status": "открыт",
    }

    # Границы окна (для поиска ближайших)
    window_min = min(start_piket, end_piket)
    window_max = max(start_piket, end_piket)

    # Реальные границы очага по ДТП (для статуса «закрыт»)
    if dtp_pk_min is not None and dtp_pk_max is not None:
        cluster_min = min(dtp_pk_min, dtp_pk_max)
        cluster_max = max(dtp_pk_min, dtp_pk_max)
    else:
        # Фоллбэк: если у ДТП нет пикетажа — используем окно
        cluster_min = window_min
        cluster_max = window_max

    # Радиус поиска ближайших (по пикетажу, в км) — от границ ДТП
    near_radius = NEAR_ROAD_PK_NP_KM if is_settlement else NEAR_ROAD_PK_OUT_KM

    # Фильтруем камеры на той же дороге
    road_cameras = [c for c in cameras if roads_match(road_name, c)]

    in_cluster_cams = []
    near_cams = []

    for cam in road_cameras:
        cam_pk = cam.get("piket")
        if cam_pk is None:
            continue

        if cluster_min <= cam_pk <= cluster_max:
            # Камера ВНУТРИ реальных границ очага (между крайними ДТП)
            in_cluster_cams.append(cam)
        elif cam_pk < cluster_min:
            # Камера до начала очага — считаем расстояние от границы ДТП
            dist_km = cluster_min - cam_pk
            if dist_km <= near_radius:
                near_cams.append((cam, dist_km))
        elif cam_pk > cluster_max:
            # Камера после конца очага — считаем расстояние от границы ДТП
            dist_km = cam_pk - cluster_max
            if dist_km <= near_radius:
                near_cams.append((cam, dist_km))

    # Камера в очаге
    if in_cluster_cams:
        result["in_cluster"] = in_cluster_cams[0]
        result["status"] = "закрыт"
        logger.debug(
            f"Закрыт очаг {road_name} [{cluster_min:.3f}..{cluster_max:.3f}] "
            f"камера {in_cluster_cams[0].get('number', '?')} "
            f"пикетаж {in_cluster_cams[0].get('piket', '?'):.3f}"
        )

    # Ближайшая камера (если нет в очаге)
    if not result["in_cluster"] and near_cams:
        near_cams.sort(key=lambda x: x[1])
        result["nearest"] = near_cams[0][0]
        result["nearest_dist_m"] = round(near_cams[0][1] * 1000)

    return result


def _search_by_gps(
    first_coords: Optional[tuple],
    last_coords: Optional[tuple],
    cameras: list[dict],
) -> dict:
    """Поиск камер по GPS-расстоянию (для очагов без пикетажа)."""
    result = {
        "in_cluster": None,
        "nearest": None,
        "nearest_dist_m": None,
        "status": "открыт",
    }

    # Точки очага (для расчёта расстояния используем обе крайние точки)
    ref_points = []
    if first_coords:
        ref_points.append(first_coords)
    if last_coords:
        ref_points.append(last_coords)
    if not ref_points:
        return result

    close_cams = []   # 200 м
    near_cams = []    # 500 м

    for cam in cameras:
        cam_lat = cam["lat"]
        cam_lon = cam["lon"]

        # Минимальное расстояние от камеры до любой точки очага
        min_dist_km = min(
            haversine_km(p[0], p[1], cam_lat, cam_lon)
            for p in ref_points
        )

        if min_dist_km <= CLOSE_RADIUS_KM:
            close_cams.append((cam, min_dist_km))
        elif min_dist_km <= NEAR_RADIUS_KM:
            near_cams.append((cam, min_dist_km))

    # Камера в очаге (200 м)
    if close_cams:
        close_cams.sort(key=lambda x: x[1])
        result["in_cluster"] = close_cams[0][0]
        result["nearest_dist_m"] = round(close_cams[0][1] * 1000)
        result["status"] = "закрыт"
        # Также запоминаем ближайшую из дальних (если есть ближе 500м)
        if near_cams:
            near_cams.sort(key=lambda x: x[1])
            if near_cams[0][1] < close_cams[0][1]:
                pass  # не перезаписываем, close имеет приоритет
    elif near_cams:
        near_cams.sort(key=lambda x: x[1])
        result["nearest"] = near_cams[0][0]
        result["nearest_dist_m"] = round(near_cams[0][1] * 1000)

    return result


def format_camera_info(cam: dict) -> str:
    """Форматирует данные камеры для ячейки Excel."""
    if not cam:
        return ""
    parts = [
        cam.get("number", ""),
        cam.get("address", ""),
    ]
    return " | ".join(p for p in parts if p)


def format_camera_coords(cam: dict) -> str:
    """Форматирует координаты камеры."""
    if not cam:
        return ""
    return f"{cam['lat']:.6f}, {cam['lon']:.6f}"


def _parse_from_dataframe(df) -> list[dict]:
    """Парсинг камер из pandas DataFrame (фоллбэк если openpyxl не справился).

    Используется когда файл в формате .xls или повреждённый .xlsx.
    """
    cameras: list[dict] = []
    # Пропускаем первые 4 строки (заголовки), данные с 5-й (индекс 4)
    for idx, row in df.iloc[4:].iterrows():
        vals = row.tolist()
        if len(vals) < 7:
            continue
        if not vals[0] or vals[2] is None:
            continue

        try:
            lat = float(vals[4]) if vals[4] else None
            lon = float(vals[5]) if vals[5] else None
        except (ValueError, TypeError):
            lat = lon = None

        if lat is None or lon is None:
            continue

        address = str(vals[6]).strip() if vals[6] else ""

        road_num, road_name, road_simple, piket, has_piket = (
            _parse_camera_address(address)
        )

        cameras.append({
            "id": str(vals[1]) if vals[1] else "",
            "number": str(vals[2]).strip(),
            "model": str(vals[3]).strip() if vals[3] else "",
            "lat": lat,
            "lon": lon,
            "address": address,
            "road_num": road_num,
            "road_name": road_name,
            "road_simple": road_simple,
            "piket": piket,
            "has_piket": has_piket,
        })

    with_piket = sum(1 for c in cameras if c["has_piket"])
    logger.info(
        f"Загружено {len(cameras)} камер "
        f"(с пикетажем: {with_piket}, городских: {len(cameras) - with_piket})"
    )
    return cameras
