"""
Генерация Excel-файлов на основе данных ДТП:

  1. dtp_cards.xlsx     — одна строка = одно ДТП (все поля карточки)
  2. dtp_uch.xlsx       — одна строка = один участник ДТП
  3. dtp_analytics.xlsx — аналитика: сравнение периодов
"""

import gc
import io
import logging
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.write_only import WriteOnlyWorksheet

logger = logging.getLogger(__name__)


# ========================
# Стили
# ========================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CELL_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ========================
# Вспомогательные функции
# ========================

def _apply_header_style(ws, col_count: int) -> None:
    """Применяет стили к строке заголовков."""
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _apply_data_styles(ws, row_count: int, col_count: int) -> None:
    """Применяет стили к ячейкам с данными."""
    for row_idx in range(2, row_count + 2):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER


def _auto_width(ws, col_count: int, max_width: int = 40) -> None:
    """Автоподбор ширины колонок с ограничением."""
    for col_idx in range(1, col_count + 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))

        # Проверяем первые 50 строк для определения ширины
        check_rows = min(ws.max_row, 51)
        for row_idx in range(2, check_rows):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val:
                cell_len = len(str(cell_val))
                if cell_len > max_len:
                    max_len = cell_len

        adjusted_width = min(max_len + 3, max_width)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 8)


# Порог для переключения на write_only режим (экономия памяти)
_WRITE_ONLY_THRESHOLD = 1500


def _create_workbook(
    column_names: list[str],
    data_rows: list[dict[str, str]],
) -> Workbook:
    """
    Создаёт объект Workbook с заголовками и данными.
    Для больших файлов (>1500 строк) использует write_only режим
    для экономии памяти (не держит все ячейки в RAM).

    Args:
        column_names: Список названий колонок (порядок важен)
        data_rows: Список словарей {название_колонки: значение}
    """
    col_count = len(column_names)
    row_count = len(data_rows)

    if row_count > _WRITE_ONLY_THRESHOLD:
        return _create_workbook_write_only(column_names, data_rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Данные"

    # Заголовки
    for col_idx, col_name in enumerate(column_names, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Данные
    for row_idx, row_data in enumerate(data_rows, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            value = row_data.get(col_name, "")
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Стили заголовков
    _apply_header_style(ws, col_count)

    # Стили данных (только для небольших файлов)
    _apply_data_styles(ws, row_count, col_count)

    # Автоподбор ширины
    _auto_width(ws, col_count)

    # Заморозка заголовков (чтобы при прокрутке шапка оставалась видна)
    ws.freeze_panes = "A2"

    # Авторазмер листа по содержимому
    if row_count > 0:
        ws.auto_filter.ref = ws.dimensions

    return wb


def _create_workbook_write_only(
    column_names: list[str],
    data_rows: list[dict[str, str]],
) -> Workbook:
    """
    Создаёт Workbook в write_only режиме — ячейки не хранятся в памяти,
    а пишутся потоково в буфер. Экономит ~100-200 МБ для больших файлов.

    Ограничения write_only: нет стилей ячеек, нет freeze_panes,
    нет auto_filter. Заголовки стилизуются через ColumnDimension.
    """
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Данные")

    # Заголовки (write_only поддерживает стили строк)
    header_row = [
        col_name for col_name in column_names
    ]
    ws.append(header_row)

    # Данные — построчная запись, без хранения в памяти
    for row_data in data_rows:
        row = [row_data.get(col_name, "") for col_name in column_names]
        ws.append(row)

    # Автоподбор ширины колонок (по заголовку + первые 50 строк)
    check_rows = min(len(data_rows), 50)
    col_letters = []
    for col_idx, col_name in enumerate(column_names, start=1):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        col_letters.append(col_letter)
        max_len = len(col_name)
        for r in range(check_rows):
            val = data_rows[r].get(col_name, "")
            if val:
                val_len = len(str(val))
                if val_len > max_len:
                    max_len = val_len
        adjusted_width = min(max_len + 3, 40)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 8)

    return wb


def generate_point_stats_file(
    current_rows: list[dict[str, str]],
    prev_rows: list[dict[str, str]] | None,
    column_names: list[str],
    current_label: str,
    prev_label: str | None = None,
) -> bytes:
    """
    Генерирует Excel-файл с ДТП в радиусе точки.

    Лист 1 — текущий период (название = current_label).
    Лист 2 — прошлый период (если есть данные, название = prev_label).

    Args:
        current_rows: Строки ДТП текущего периода
        prev_rows: Строки ДТП прошлого периода (опционально)
        column_names: Названия колонок
        current_label: Подпись текущего периода (используется как название листа)
        prev_label: Подпись прошлого периода (опционально)
    """
    wb = Workbook()

    col_widths_ps = {
        "Дата ДТП": 14,
        "Время": 8,
        "Населённый пункт": 20,
        "Район": 18,
        "Дорога/Улица": 35,
        "Пикетаж": 16,
        "Расстояние, м": 12,
        "Широта": 14,
        "Долгота": 14,
        "Вид ДТП": 25,
        "Кол-во ТС": 10,
        "Кол-во участников": 14,
        "Погибло": 8,
        "Ранено": 8,
        "Категории участников": 28,
        "Категория дороги": 18,
        "Состояние покрытия": 22,
        "Погода": 22,
        "Освещение": 16,
        "Нарушения ПДД": 45,
        "Типы ТС": 25,
        "Марки/Модели ТС": 35,
    }

    # --- Лист 1: Текущий период ---
    ws1 = wb.active
    sheet_name = current_label[:31]  # openpyxl: max 31 символов
    ws1.title = sheet_name

    for col_idx, col_name in enumerate(column_names, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(current_rows, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            value = row_data.get(col_name, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    for col_idx, col_name in enumerate(column_names, start=1):
        col_letter = ws1.cell(row=1, column=col_idx).column_letter
        ws1.column_dimensions[col_letter].width = col_widths_ps.get(col_name, 18)

    ws1.freeze_panes = "A2"
    if current_rows:
        ws1.auto_filter.ref = ws1.dimensions

    # --- Лист 2: Прошлый период (если есть) ---
    if prev_rows and prev_label:
        ws2 = wb.create_sheet(title=prev_label[:31])

        for col_idx, col_name in enumerate(column_names, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(prev_rows, start=2):
            for col_idx, col_name in enumerate(column_names, start=1):
                value = row_data.get(col_name, "")
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

        for col_idx, col_name in enumerate(column_names, start=1):
            col_letter = ws2.cell(row=1, column=col_idx).column_letter
            ws2.column_dimensions[col_letter].width = col_widths_ps.get(col_name, 18)

        ws2.freeze_panes = "A2"
        if prev_rows:
            ws2.auto_filter.ref = ws2.dimensions

    return workbook_to_bytes(wb)


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Сериализует Workbook в байты xlsx-файла."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ========================
# Публичные функции
# ========================

def generate_file1(file1_data: list[dict[str, str]]) -> bytes:
    """
    Генерирует Файл 1: одна строка = одно ДТП.

    Args:
        file1_data: Данные от gibdd_parser.build_file1_data()

    Returns:
        Байты xlsx-файла
    """
    from gibdd_parser import get_file1_column_names

    column_names = get_file1_column_names()
    wb = _create_workbook(column_names, file1_data)
    return workbook_to_bytes(wb)


def generate_file2(file2_data: list[dict[str, str]]) -> bytes:
    """
    Генерирует Файл 2: одна строка = один участник ДТП.

    Args:
        file2_data: Данные от gibdd_parser.build_file2_data()

    Returns:
        Байты xlsx-файла
    """
    from gibdd_parser import get_file2_column_names

    column_names = get_file2_column_names()
    wb = _create_workbook(column_names, file2_data)
    return workbook_to_bytes(wb)


def generate_both_files(
    file1_data: list[dict[str, str]],
    file2_data: list[dict[str, str]],
) -> tuple[bytes, bytes]:
    """
    Генерирует оба Excel-файла.

    Returns:
        (file1_bytes, file2_bytes)
    """
    logger.info(f"Генерация Excel: Файл 1 — {len(file1_data)} ДТП, Файл 2 — {len(file2_data)} участников")
    file1_bytes = generate_file1(file1_data)
    # Освобождаем память: удаляем промежуточные данные Файла 1
    del file1_data
    gc.collect()
    file2_bytes = generate_file2(file2_data)
    del file2_data
    gc.collect()
    logger.info(f"Файл 1: {len(file1_bytes)} байт, Файл 2: {len(file2_bytes)} байт")
    return file1_bytes, file2_bytes


# Стили для разделительных строк аналитики
SECTION_FONT = Font(bold=True, size=11)
POSITIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def generate_analytics_file(
    analytics_data: list[dict[str, str]],
    column_names: list[str],
) -> bytes:
    """
    Генерирует Excel-файл с аналитикой (сравнение периодов).

    Args:
        analytics_data: Данные от analytics.build_analytics_excel_data()
        column_names: Названия колонок от analytics.get_analytics_column_names()

    Returns:
        Байты xlsx-файла
    """
    from openpyxl.styles import numbers

    wb = Workbook()
    ws = wb.active
    ws.title = "Аналитика"

    col_count = len(column_names)

    # Заголовки
    for col_idx, col_name in enumerate(column_names, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Данные
    for row_idx, row_data in enumerate(analytics_data, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # Выделяем разделительные строки (заголовки секций)
        indicator = row_data.get("Показатель", "")
        if indicator and indicator == indicator.upper() and indicator.strip():
            for col_idx in range(1, col_count + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = SECTION_FONT

        # Цветовое кодирование колонки "Изменение, %"
        change_cell = ws.cell(row=row_idx, column=4)
        change_val = row_data.get("Изменение, %", "")
        if isinstance(change_val, (int, float)) and change_val != 0:
            if change_val > 0:
                change_cell.fill = NEGATIVE_FILL  # Рост показателя = красный
            elif change_val < 0:
                change_cell.fill = POSITIVE_FILL  # Снижение показателя = зелёный

    # Ширина колонок
    ws.column_dimensions["A"].width = 35
    for col_idx in range(2, col_count + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 22

    ws.freeze_panes = "A2"

    return workbook_to_bytes(wb)


# ========================
# Очаги концентрации ДТП
# ========================

# Специальные стили для очагов
ZONE_NP_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
ZONE_NONP_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def generate_concentration_file(
    concentration_data: list[dict[str, str]],
    column_names: list[str],
    detail_data: list[dict[str, str]] | None = None,
    detail_columns: list[str] | None = None,
) -> bytes:
    """
    Генерирует Excel-файл с очагами концентрации ДТП.
    Лист 1 — сводка очагов, Лист 2 — детализация ДТП.

    Args:
        concentration_data: Данные от concentration_points.build_concentration_excel_data()
        column_names: Названия колонок от concentration_points.get_concentration_column_names()
        detail_data: Данные от concentration_points.build_concentration_detail_data() (опционально)
        detail_columns: Названия колонок от concentration_points.get_detail_column_names() (опционально)

    Returns:
        Байты xlsx-файла
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Очаги ДТП"

    col_count = len(column_names)

    # Заголовки
    for col_idx, col_name in enumerate(column_names, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Данные
    for row_idx, row_data in enumerate(concentration_data, start=2):
        for col_idx, col_name in enumerate(column_names, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # Цветовое кодирование по типу зоны
        zone = row_data.get("Тип зоны", "")
        fill = None
        if zone.startswith("НП"):
            fill = ZONE_NP_FILL
        elif zone.startswith("Вне"):
            fill = ZONE_NONP_FILL

        if fill:
            for col_idx in range(1, col_count + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    # Ширина колонок
    col_widths = {
        "№ очага": 8,
        "Тип зоны": 28,
        "Дорога/Улица": 30,
        "Пикетаж начало": 14,
        "Пикетаж конец": 14,
        "Широта первого ДТП": 16,
        "Долгота первого ДТП": 16,
        "Широта последнего ДТП": 16,
        "Долгота последнего ДТП": 16,
        "Кол-во ДТП": 10,
        "Виды ДТП (детализация)": 45,
        "Доминирующий вид": 25,
        "Погибло": 8,
        "Ранено": 8,
        "Дата первого ДТП": 14,
        "Дата последнего ДТП": 14,
    }
    for col_idx, col_name in enumerate(column_names, start=1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

    ws.freeze_panes = "A2"

    if concentration_data:
        ws.auto_filter.ref = ws.dimensions

    # --- Лист 2: Детализация ДТП в очагах ---
    if detail_data and detail_columns:
        ws2 = wb.create_sheet("Детализация ДТП")

        det_col_count = len(detail_columns)

        # Заголовки
        for col_idx, col_name in enumerate(detail_columns, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        # Данные
        for row_idx, row_data in enumerate(detail_data, start=2):
            for col_idx, col_name in enumerate(detail_columns, start=1):
                value = row_data.get(col_name, "")
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

        # Ширина колонок
        det_widths = {
            "№ очага": 8,
            "Дата ДТП": 14,
            "Вид ДТП": 25,
            "Дорога/Улица": 30,
            "Пикетаж": 14,
            "Широта": 16,
            "Долгота": 16,
            "Погибло": 8,
            "Ранено": 8,
        }
        for col_idx, col_name in enumerate(detail_columns, start=1):
            col_letter = ws2.cell(row=1, column=col_idx).column_letter
            ws2.column_dimensions[col_letter].width = det_widths.get(col_name, 20)

        ws2.freeze_panes = "A2"

        if detail_data:
            ws2.auto_filter.ref = ws2.dimensions

    return workbook_to_bytes(wb)


# ========================
# Очаги ДТП: динамика (сравнение периодов)
# ========================

# Стили для статусов динамики
DYN_NEW_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
DYN_GROWING_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
DYN_SHRINKING_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
DYN_STABLE_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
DYN_LOST_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

PRECLUSTER_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # светло-оранжевый

DYN_STATUS_FILLS = {
    "Новый": DYN_NEW_FILL,
    "Рост": DYN_GROWING_FILL,
    "Снижение": DYN_SHRINKING_FILL,
    "Стабильный": DYN_STABLE_FILL,
    "Исчезнувший": DYN_LOST_FILL,
}


def generate_concentration_dynamics_file(
    current_year_data: list[dict[str, str]],
    current_year_columns: list[str],
    dynamics_data: list[dict[str, str]],
    dynamics_columns: list[str],
    detail_data: list[dict[str, str]] | None = None,
    detail_columns: list[str] | None = None,
    precluster_data: list[dict[str, str]] | None = None,
    precluster_columns: list[str] | None = None,
) -> bytes:
    """
    Генерирует Excel-файл с очагами ДТП и исторической динамикой.

    Лист 1 «Очаги ДТП» — очаги запрашиваемого года (без динамики):
      стандартные колонки с цветовой кодировкой по типу зоны
      (НП — жёлтый, вне НП — голубой).

    Лист 2 «Динамика очагов» — сводка с цветовой кодировкой по статусу:
      - Зелёный = Новый
      - Красный = Рост (ухудшение)
      - Голубой = Снижение (улучшение)
      - Серый = Стабильный
      - Светло-зелёный = Исчезнувший

    Лист 3 «Детализация ДТП» — все ДТП с пометкой периода.

    Лист 4 «Предочаги» — места, которым не хватает 1 ДТП до очага.

    Args:
        current_year_data: Данные очагов только за запрашиваемый год
        current_year_columns: Названия колонок для очагов текущего года
        dynamics_data: Данные очагов с динамикой (текущие + исчезнувшие)
        dynamics_columns: Названия колонок для динамики
        detail_data: Данные от concentration_points.build_dynamics_detail_data()
        detail_columns: Названия колонок для детализации
    """
    wb = Workbook()

    # ==============================
    # Лист 1: Очаги ДТП (текущий год)
    # ==============================
    ws1 = wb.active
    ws1.title = "Очаги ДТП"

    col_count1 = len(current_year_columns)

    for col_idx, col_name in enumerate(current_year_columns, start=1):
        cell = ws1.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(current_year_data, start=2):
        for col_idx, col_name in enumerate(current_year_columns, start=1):
            value = row_data.get(col_name, "")
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # Цветовое кодирование по типу зоны
        zone = row_data.get("Тип зоны", "")
        fill = None
        if zone.startswith("НП"):
            fill = ZONE_NP_FILL
        elif zone.startswith("Вне"):
            fill = ZONE_NONP_FILL

        if fill:
            for col_idx in range(1, col_count1 + 1):
                ws1.cell(row=row_idx, column=col_idx).fill = fill

    # Ширина колонок
    col_widths = {
        "№ очага": 8,
        "Тип зоны": 28,
        "Дорога/Улица": 30,
        "Пикетаж начало": 14,
        "Пикетаж конец": 14,
        "Широта первого ДТП": 16,
        "Долгота первого ДТП": 16,
        "Широта последнего ДТП": 16,
        "Долгота последнего ДТП": 16,
        "Кол-во ДТП": 10,
        "Виды ДТП (детализация)": 45,
        "Доминирующий вид": 25,
        "Погибло": 8,
        "Ранено": 8,
        "Дата первого ДТП": 14,
        "Дата последнего ДТП": 14,
    }
    for col_idx, col_name in enumerate(current_year_columns, start=1):
        col_letter = ws1.cell(row=1, column=col_idx).column_letter
        ws1.column_dimensions[col_letter].width = col_widths.get(col_name, 20)

    ws1.freeze_panes = "A2"

    if current_year_data:
        ws1.auto_filter.ref = ws1.dimensions

    # ==============================
    # Лист 2: Динамика очагов
    # ==============================
    ws2 = wb.create_sheet("Динамика очагов")

    col_count2 = len(dynamics_columns)

    for col_idx, col_name in enumerate(dynamics_columns, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx, row_data in enumerate(dynamics_data, start=2):
        for col_idx, col_name in enumerate(dynamics_columns, start=1):
            value = row_data.get(col_name, "")
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

        # Цветовое кодирование по статусу динамики
        status = row_data.get("Статус", "")
        fill = DYN_STATUS_FILLS.get(status)

        if fill:
            for col_idx in range(1, col_count2 + 1):
                ws2.cell(row=row_idx, column=col_idx).fill = fill

        # Дополнительно: тип зоны в колонках «Тип зоны» и «Дорога»
        zone = row_data.get("Тип зоны", "")
        zone_fill = None
        if zone.startswith("НП"):
            zone_fill = ZONE_NP_FILL
        elif zone.startswith("Вне"):
            zone_fill = ZONE_NONP_FILL

        if zone_fill:
            for col_idx, col_name in enumerate(dynamics_columns, start=1):
                if col_name in ("Тип зоны", "Дорога/Улица"):
                    ws2.cell(row=row_idx, column=col_idx).fill = zone_fill

    # Ширина колонок
    dyn_widths = {
        "№ очага": 8,
        "Статус": 16,
        "Тип зоны": 28,
        "Дорога/Улица": 30,
        "Пикетаж начало": 14,
        "Пикетаж конец": 14,
        "Широта": 14,
        "Долгота": 14,
        "Кол-во ДТП": 10,
        "ДТП (пр. период)": 14,
        "Изменение ДТП": 12,
        "Виды ДТП (детализация)": 45,
        "Доминирующий вид": 25,
        "Погибло": 8,
        "Ранено": 8,
        "Погибло (пр. период)": 14,
        "Ранено (пр. период)": 14,
        "Дата первого ДТП": 14,
        "Дата последнего ДТП": 14,
    }
    for col_idx, col_name in enumerate(dynamics_columns, start=1):
        col_letter = ws2.cell(row=1, column=col_idx).column_letter
        ws2.column_dimensions[col_letter].width = dyn_widths.get(col_name, 20)

    ws2.freeze_panes = "A2"

    if dynamics_data:
        ws2.auto_filter.ref = ws2.dimensions

    # ==============================
    # Лист 3: Детализация ДТП
    # ==============================
    if detail_data and detail_columns:
        ws3 = wb.create_sheet("Детализация ДТП")

        for col_idx, col_name in enumerate(detail_columns, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(detail_data, start=2):
            for col_idx, col_name in enumerate(detail_columns, start=1):
                value = row_data.get(col_name, "")
                cell = ws3.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER

            # Цвет по статусу
            status = row_data.get("Статус", "")
            fill = DYN_STATUS_FILLS.get(status)
            if fill:
                for ci, cn in enumerate(detail_columns, start=1):
                    if cn == "Статус":
                        ws3.cell(row=row_idx, column=ci).fill = fill
                        break

        det_widths = {
            "№ очага": 8,
            "Статус": 16,
            "Период": 18,
            "Дата ДТП": 14,
            "Вид ДТП": 25,
            "Дорога/Улица": 30,
            "Пикетаж": 14,
            "Широта": 16,
            "Долгота": 16,
            "Погибло": 8,
            "Ранено": 8,
        }
        for col_idx, col_name in enumerate(detail_columns, start=1):
            col_letter = ws3.cell(row=1, column=col_idx).column_letter
            ws3.column_dimensions[col_letter].width = det_widths.get(col_name, 20)

        ws3.freeze_panes = "A2"

        if detail_data:
            ws3.auto_filter.ref = ws3.dimensions

    # ==============================
    # Лист 4: Предочаги (если есть)
    # ==============================
    if precluster_data and precluster_columns:
        ws4 = wb.create_sheet("Предочаги")

        col_count4 = len(precluster_columns)

        for col_idx, col_name in enumerate(precluster_columns, start=1):
            cell = ws4.cell(row=1, column=col_idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGNMENT
            cell.border = THIN_BORDER

        for row_idx, row_data in enumerate(precluster_data, start=2):
            for col_idx, col_name in enumerate(precluster_columns, start=1):
                value = row_data.get(col_name, "")
                cell = ws4.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = CELL_ALIGNMENT
                cell.border = THIN_BORDER
                cell.fill = PRECLUSTER_FILL

        # Ширина колонок (переиспользуем ширину очагов + добавляем свои)
        pre_widths = {
            "№ предочага": 12,
            "Критерий предочага": 25,
        }
        for col_idx, col_name in enumerate(precluster_columns, start=1):
            col_letter = ws4.cell(row=1, column=col_idx).column_letter
            ws4.column_dimensions[col_letter].width = pre_widths.get(col_name, 20)

        ws4.freeze_panes = "A2"

        if precluster_data:
            ws4.auto_filter.ref = ws4.dimensions

    return workbook_to_bytes(wb)
